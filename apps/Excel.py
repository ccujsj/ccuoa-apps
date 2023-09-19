from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import os
import win32com.client as win32  # 将Win32com导入为win32
from .Logger import logger
import pandas as pd


class Excel:
    """
    more for:https://www.yuque.com/yhwh/cs/wol6kz
    """

    def __init__(self, filename: str):
        self.filename = filename
        self.wb = None
        try:
            self.wb: Workbook = load_workbook(filename)
        except InvalidFileException as e:
            logger.debug(e)
            logger.info("converting to xlsx")
            current_working_directory = os.getcwd()
            xls_filename = os.path.join(current_working_directory, filename)  # 获取绝对地址文件名
            excel = win32.gencache.EnsureDispatch('Excel.Application')  # 调用win32
            wb = excel.Workbooks.Open(xls_filename)  # 存为wb对象
            xlsx_filename = xls_filename + "x"
            wb.SaveAs(xlsx_filename, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls  extension
            excel.Application.Quit()
            self.wb = load_workbook(xlsx_filename)
            logger.info(f"{filename} has converted to xlsx file {xlsx_filename}")

        self.active_worksheet = self.wb.active

    @property
    def props(self):
        """
        dimensions	表格的大小，这里的大小是指含有数据的表格的大小，即：左上角的坐标:右下角的坐标
        max_row	表格的最大行
        min_row	表格的最小行
        max_column	表格的最大列
        min_column	表格的最小列
        rows	按行获取单元格(Cell对象) - 生成器           工作表.rows
        columns	按列获取单元格(Cell对象) - 生成器           工作表.columns
        freeze_panes	冻结窗格            工作表.freeze_panes = "C3"
        values	按行获取表格的内容(数据)  - 生成器
        """
        return self.active_worksheet

    def get_header_mapping(self):
        head_line = None
        for row in self.active_worksheet.iter_rows():
            head_line = row  # return a multi-element tuple(<A1>,<A2>,...)
            break
        mapdict = {}
        row = 1
        fields = list(map(lambda x: x.value, list(head_line)))  # travel list [ <A1>.value,<A2>.value,... ]
        if self.__t_check_list_duplication(fields):
            logger.warning(
                "field line:" + str(fields) + str(" are duplicated, It might be interpreted in unexpected ways"))
        for item in fields:
            mapdict.update({item: row})  # create mapping {"field_name":colum_number}
            row += 1
        return mapdict

    def get_templates_render_dicts(self) -> list:
        mapping = self.get_header_mapping()
        lists = []
        for i in self.get_ved():
            new_mapping = mapping.copy()
            for items in mapping.items():
                key = items[0]       # jinjia2 template variables
                val = int(items[1])  # position
                new_mapping.update({key: i[val - 1]})
            lists.append(new_mapping)
        return lists

    @staticmethod
    def __t_check_list_duplication(target_list: list):
        tool_dict = {}
        idx = 0
        duplications = []
        for item in target_list:
            try:
                idx = tool_dict[item]
                duplications.append(item)
            except KeyError:
                tool_dict.update({item: idx})
                idx += 1
        if len(duplications) == 0:
            return None
        return duplications

    def get_ved(self):
        lines = 0
        table = []
        for row in self.active_worksheet.iter_rows():
            if lines == 0:
                lines += 1  # skip header
                continue
            else:
                table.append(list(map(lambda x: x.value, list(row))))
        # table should be a two-dimension value array
        return table

    def get_pd_dataframe(self):
        return pd.DataFrame(self.get_ved())
